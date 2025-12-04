"""
Excel I/O Service for Market Brosur Sistemi
Handles Excel parsing and export with Turkish number format support
"""

import openpyxl
from openpyxl import Workbook
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter
import re
from typing import Dict, List, Tuple, Optional


def parse_turkish_float(value: str) -> float:
    """
    Parse Turkish number format (1.234,56) to float
    Also supports international format (1234.56)
    
    Args:
        value: String number in Turkish or international format
        
    Returns:
        Parsed float value
        
    Raises:
        ValueError: If value cannot be parsed
    """
    if not value or not isinstance(value, str):
        return 0.0
    
    value = value.strip()
    if not value:
        return 0.0
    
    # Check if Turkish format (contains comma as decimal separator)
    if ',' in value:
        # Turkish format: 1.234,56 -> remove thousands separator (.), replace comma with dot
        value = value.replace('.', '').replace(',', '.')
    
    try:
        return float(value)
    except ValueError:
        raise ValueError(f"Cannot parse number: {value}")


def validate_excel_schema(headers: List[str]) -> Tuple[bool, Optional[str]]:
    """
    Validate Excel column headers match required schema
    
    Required headers:
    - Barkod (required)
    - Ürün Adı (optional, ASCII fallback: Urun Adi)
    - İndirim Fiyatı (required, ASCII fallback: Indirim Fiyati)
    - Normal Fiyat (optional, ASCII fallback: Normal Fiyat)
    
    Args:
        headers: List of column headers from Excel file
        
    Returns:
        Tuple of (is_valid, error_message)
    """
    # Normalize headers (strip whitespace, case-insensitive)
    normalized = [h.strip() for h in headers if h]
    
    # Required headers with ASCII fallbacks
    required_headers = {
        'Barkod': ['Barkod', 'barkod'],
        'İndirim Fiyatı': ['İndirim Fiyatı', 'Indirim Fiyati', 'indirim fiyati', 'İndirim Fiyati'],
    }
    
    # Check for required headers
    for header_name, variants in required_headers.items():
        if not any(variant in normalized for variant in variants):
            return False, f"Missing required column: {header_name}"
    
    return True, None


def parse_excel_file(file_path: str, max_rows: int = 5000) -> Tuple[List[Dict], List[str], Dict]:
    """
    Parse Excel file and extract product data
    
    Args:
        file_path: Path to Excel file
        max_rows: Maximum number of rows to process (security limit)
        
    Returns:
        Tuple of (products_list, validation_errors, stats)
        
    Security:
        - Only reads first sheet
        - Strips formulas (data_only=True)
        - Enforces row limit
        - Validates schema before processing
    """
    validation_errors = []
    products = []
    stats = {
        'total_rows': 0,
        'processed': 0,
        'duplicates_skipped': 0,
        'errors': 0
    }
    
    try:
        # Load workbook (data_only=True strips formulas for security)
        wb = openpyxl.load_workbook(file_path, data_only=True, read_only=True)
        
        # Only use first sheet
        sheet = wb.worksheets[0]
        
        # Get headers from first row
        headers = []
        for cell in sheet[1]:
            headers.append(str(cell.value).strip() if cell.value else '')
        
        # Validate schema
        is_valid, error_msg = validate_excel_schema(headers)
        if not is_valid:
            validation_errors.append(f"Schema validation failed: {error_msg}")
            return products, validation_errors, stats
        
        # Create header mapping (handle ASCII fallbacks)
        header_map = {}
        for idx, header in enumerate(headers):
            h = header.strip()
            if h in ['Barkod', 'barkod']:
                header_map['Barkod'] = idx
            elif h in ['Ürün Adı', 'Urun Adi', 'urun adi']:
                header_map['Ürün Adı'] = idx
            elif h in ['İndirim Fiyatı', 'Indirim Fiyati', 'indirim fiyati', 'İndirim Fiyati']:
                header_map['İndirim Fiyatı'] = idx
            elif h in ['Normal Fiyat', 'normal fiyat']:
                header_map['Normal Fiyat'] = idx
        
        # Track existing barcodes for duplicate detection
        existing_barcodes = set()
        
        # Process data rows (skip header row)
        row_num = 1
        for row in sheet.iter_rows(min_row=2, values_only=True):
            row_num += 1
            
            # Security: enforce row limit
            if row_num > max_rows:
                validation_errors.append(f"File too large: maximum {max_rows} rows allowed")
                stats['errors'] += 1
                wb.close()
                return products, validation_errors, stats
            
            stats['total_rows'] += 1
            
            # Skip empty rows
            if not any(row):
                continue
            
            try:
                # Extract barcode
                barcode_idx = header_map.get('Barkod')
                if barcode_idx is None or barcode_idx >= len(row):
                    validation_errors.append(f"Row {row_num}: Missing barcode column")
                    stats['errors'] += 1
                    continue
                
                barcode = str(row[barcode_idx]).strip() if row[barcode_idx] else ''
                if not barcode:
                    validation_errors.append(f"Row {row_num}: Barcode cannot be empty")
                    stats['errors'] += 1
                    continue
                
                # Check for duplicates
                if barcode in existing_barcodes:
                    stats['duplicates_skipped'] += 1
                    continue
                
                existing_barcodes.add(barcode)
                
                # Extract product name (optional)
                name_idx = header_map.get('Ürün Adı')
                name = ''
                if name_idx is not None and name_idx < len(row) and row[name_idx]:
                    name = str(row[name_idx]).strip()
                
                # Extract discount price (required)
                discount_idx = header_map.get('İndirim Fiyatı')
                if discount_idx is None or discount_idx >= len(row):
                    validation_errors.append(f"Row {row_num}: Missing discount price column")
                    stats['errors'] += 1
                    continue
                
                discount_price = 0.0
                discount_value = row[discount_idx]
                if discount_value:
                    if isinstance(discount_value, (int, float)):
                        discount_price = float(discount_value)
                    else:
                        discount_price = parse_turkish_float(str(discount_value))
                
                if discount_price <= 0:
                    validation_errors.append(f"Row {row_num}: Discount price must be greater than 0")
                    stats['errors'] += 1
                    continue
                
                # Extract normal price (optional)
                normal_price = 0.0
                normal_idx = header_map.get('Normal Fiyat')
                if normal_idx is not None and normal_idx < len(row) and row[normal_idx]:
                    normal_value = row[normal_idx]
                    if isinstance(normal_value, (int, float)):
                        normal_price = float(normal_value)
                    else:
                        try:
                            normal_price = parse_turkish_float(str(normal_value))
                        except ValueError:
                            # Ignore invalid normal price (it's optional)
                            normal_price = 0.0
                
                # Create product entry
                product = {
                    'barcode': barcode,
                    'name': name,
                    'discount_price': discount_price,
                    'normal_price': normal_price,
                    'product_group': ''  # Will be filled by CAMGOZ API or manual entry
                }
                
                products.append(product)
                stats['processed'] += 1
                
            except Exception as e:
                validation_errors.append(f"Row {row_num}: {str(e)}")
                stats['errors'] += 1
                continue
        
        wb.close()
        
    except Exception as e:
        validation_errors.append(f"Failed to parse Excel file: {str(e)}")
        stats['errors'] += 1
    
    return products, validation_errors, stats


def export_to_excel(products: List[Dict], output_path: str) -> bool:
    """
    Export products to Excel file
    
    Args:
        products: List of product dictionaries
        output_path: Path to save Excel file
        
    Returns:
        True if successful, False otherwise
    """
    try:
        wb = Workbook()
        ws: Worksheet = wb.active  # type: ignore
        ws.title = "Ürünler"
        
        # Set headers
        headers = ['Barkod', 'Ürün Adı', 'İndirim Fiyatı', 'Normal Fiyat']
        ws.append(headers)
        
        # Style headers (bold)
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num)
            cell.font = Font(bold=True)
        
        # Add data rows
        for product in products:
            ws.append([
                product.get('barcode', ''),
                product.get('name', ''),
                product.get('discount_price', 0.0),
                product.get('normal_price', 0.0)
            ])
        
        # Auto-adjust column widths
        for col_num, header in enumerate(headers, 1):
            ws.column_dimensions[get_column_letter(col_num)].width = 20
        
        # Save workbook
        wb.save(output_path)
        wb.close()
        
        return True
        
    except Exception as e:
        print(f"Failed to export Excel: {str(e)}")
        return False
